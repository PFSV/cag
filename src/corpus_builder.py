"""Build one deterministic text corpus from common document formats."""

from __future__ import annotations

import logging
import re
from pathlib import Path

LOGGER = logging.getLogger(__name__)


def read_text_files(data_root: Path) -> list[dict[str, str]]:
    """Read non-empty text-like files below *data_root*.

    CAG is data-source agnostic after serialization to text. TXT, Markdown,
    JSON, JSONL, and CSV are handled directly; applications can add a loader
    for PDFs, databases, or APIs and pass the resulting text through the same
    rendering path.
    """
    documents: list[dict[str, str]] = []
    for path in sorted(data_root.rglob("*")):
        if path.suffix.lower() not in {".txt", ".md", ".json", ".jsonl", ".csv"} or not path.is_file():
            continue
        text = path.read_text(encoding="utf-8", errors="ignore").strip()
        if text:
            relative = path.relative_to(data_root).as_posix()
            documents.append({"source": relative, "section": path.parts[len(data_root.parts)], "text": text})
    return documents


def read_faq(xlsx_path: Path) -> list[dict[str, str]]:
    """Read only the ``key-answer`` FAQ sheet to avoid paraphrase duplication."""
    import openpyxl

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "key-answer" not in workbook.sheetnames:
        LOGGER.warning("Skipping %s: no key-answer sheet", xlsx_path)
        return []
    documents = []
    for row in list(workbook["key-answer"].iter_rows(values_only=True))[1:]:
        if len(row) < 5:
            continue
        _, question, answer, _, category = row[:5]
        if question and answer:
            category = str(category or "uncategorized").strip()
            documents.append({
                "source": f"{xlsx_path.name}#key-answer",
                "section": f"FAQ-{category}",
                "text": f"Q: {str(question).strip()}\nA: {str(answer).strip()}",
            })
    workbook.close()
    return documents


def build_documents(data_root: Path, xlsx_path: Path | None = None) -> list[dict[str, str]]:
    """Collect raw documents, optionally adding one FAQ workbook."""
    documents = read_text_files(data_root)
    if xlsx_path is None:
        candidates = sorted(data_root.rglob("*.xlsx"))
        xlsx_path = candidates[0] if candidates else None
    if xlsx_path and xlsx_path.exists():
        documents.extend(read_faq(xlsx_path))
    return documents


def _clean(text: str) -> str:
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def render_corpus(documents: list[dict[str, str]]) -> str:
    """Serialize documents into stable section blocks for KV pre-filling."""
    sections: dict[str, list[str]] = {}
    for document in documents:
        sections.setdefault(document["section"], []).append(_clean(document["text"]))
    return "\n\n---\n\n".join(
        f"## {section}\n\n{chr(10).join(texts)}" for section, texts in sections.items()
    )


def rough_token_estimate(text: str) -> int:
    """Return a deliberately rough estimate; use the model tokenizer for limits."""
    return int(len(text) / 1.7)


def build_corpus(data_root: Path, output: Path, xlsx_path: Path | None = None) -> int:
    """Build and write the corpus, returning the number of source documents."""
    documents = build_documents(data_root, xlsx_path)
    if not documents:
        raise ValueError(f"No supported documents found below {data_root}")
    corpus = render_corpus(documents)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(corpus + "\n", encoding="utf-8")
    LOGGER.info("Wrote %d documents to %s (%d chars, ~%d tokens)", len(documents), output, len(corpus), rough_token_estimate(corpus))
    return len(documents)
